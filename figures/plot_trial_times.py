#!/usr/bin/env python

from pylab import *
from openpyxl.reader.excel import load_workbook
from collections import defaultdict

wb = load_workbook(filename = 'trial_data.xlsx')

times_by_scenario = defaultdict(list)
hits_by_scenario = defaultdict(list)
hits = []

for worksheet in wb.worksheets:
	temp = worksheet.cell('H2').value
	temp = (float(temp.hour) * 60.0) + float(temp.minute)
	times_by_scenario['camera_0_latency'].append(temp)
	hits_by_scenario['camera_0_latency'].append(worksheet.cell('I2').value)
	temp = worksheet.cell('H3').value
	temp = (float(temp.hour) * 60.0) + float(temp.minute)
	times_by_scenario['camera_1_latency'].append(temp)
	hits_by_scenario['camera_1_latency'].append(worksheet.cell('I3').value)
	temp = worksheet.cell('H4').value
	temp = (float(temp.hour) * 60.0) + float(temp.minute)
	times_by_scenario['camera_2_latency'].append(temp)
	hits_by_scenario['camera_2_latency'].append(worksheet.cell('I4').value)
	temp = worksheet.cell('H5').value
	temp = (float(temp.hour) * 60.0) + float(temp.minute)
	times_by_scenario['3d_0_latency'].append(temp)
	hits_by_scenario['3d_0_latency'].append(worksheet.cell('I5').value)
	temp = worksheet.cell('H6').value
	temp = (float(temp.hour) * 60.0) + float(temp.minute)
	times_by_scenario['3d_1_latency'].append(temp)
	hits_by_scenario['3d_1_latency'].append(worksheet.cell('I6').value)
	temp = worksheet.cell('H7').value
	temp = (float(temp.hour) * 60.0) + float(temp.minute)
	times_by_scenario['3d_2_latency'].append(temp)
	hits_by_scenario['3d_2_latency'].append(worksheet.cell('I7').value)
	temp = worksheet.cell('H8').value
	temp = (float(temp.hour) * 60.0) + float(temp.minute)
	times_by_scenario['vehicle_1_latency'].append(temp)
	hits_by_scenario['vehicle_1_latency'].append(worksheet.cell('I8').value)
	temp = worksheet.cell('H9').value
	temp = (float(temp.hour) * 60.0) + float(temp.minute)
	times_by_scenario['vehicle_2_latency'].append(temp)
	hits_by_scenario['vehicle_2_latency'].append(worksheet.cell('I9').value)


scenarios = ['camera', 'camera+3D', 'camera+3D+prediction']

avg_0 = []
avg_1 = []
avg_2 = []

avg_0.append(float(sum(times_by_scenario['camera_0_latency']))/len(times_by_scenario['camera_0_latency']))
avg_1.append(float(sum(times_by_scenario['camera_1_latency']))/len(times_by_scenario['camera_1_latency']))
avg_2.append(float(sum(times_by_scenario['camera_2_latency']))/len(times_by_scenario['camera_2_latency']))
avg_0.append(float(sum(times_by_scenario['3d_0_latency']))/len(times_by_scenario['3d_0_latency']))
avg_1.append(float(sum(times_by_scenario['3d_1_latency']))/len(times_by_scenario['3d_1_latency']))
avg_2.append(float(sum(times_by_scenario['3d_2_latency']))/len(times_by_scenario['3d_2_latency']))
avg_1.append(float(sum(times_by_scenario['vehicle_1_latency']))/len(times_by_scenario['vehicle_1_latency']))
avg_2.append(float(sum(times_by_scenario['vehicle_2_latency']))/len(times_by_scenario['vehicle_2_latency']))

ylabel('Time in seconds to complete the course')
xticks([1,2,3], scenarios)
xlim(0.5, 3.5)
p1, = plot([1,2], avg_0, 'b+-', linewidth=2)
hold(True)

p2, = plot([1,2,3], avg_1, 'r.-', linewidth=2)

p3, = plot([1,2,3], avg_2, 'gx-', linewidth=2)

title('Average Trial Times by Scenario and Latency')

legend([p1, p2, p3], ['No Latency', '1 second RTT', '2 second RTT'], loc=7)

show()

avg_0 = []
avg_1 = []
avg_2 = []

avg_0.append(float(sum(hits_by_scenario['camera_0_latency']))/len(hits_by_scenario['camera_0_latency']))
avg_1.append(float(sum(hits_by_scenario['camera_1_latency']))/len(hits_by_scenario['camera_1_latency']))
avg_2.append(float(sum(hits_by_scenario['camera_2_latency']))/len(hits_by_scenario['camera_2_latency']))
avg_0.append(float(sum(hits_by_scenario['3d_0_latency']))/len(hits_by_scenario['3d_0_latency']))
avg_1.append(float(sum(hits_by_scenario['3d_1_latency']))/len(hits_by_scenario['3d_1_latency']))
avg_2.append(float(sum(hits_by_scenario['3d_2_latency']))/len(hits_by_scenario['3d_2_latency']))
avg_1.append(float(sum(hits_by_scenario['vehicle_1_latency']))/len(hits_by_scenario['vehicle_1_latency']))
avg_2.append(float(sum(hits_by_scenario['vehicle_2_latency']))/len(hits_by_scenario['vehicle_2_latency']))

ylabel('Number of hits during the trial course')
xticks([1,2,3], scenarios)
xlim(0.5, 3.5)
p1, = plot([1,2], avg_0, 'b+-', linewidth=2)
hold(True)

p2, = plot([1,2,3], avg_1, 'r.-', linewidth=2)

p3, = plot([1,2,3], avg_2, 'gx-', linewidth=2)

title('Average Number of Hits by Scenario and Latency')

legend([p1, p2, p3], ['No Latency', '1 second RTT', '2 second RTT'], loc=1)

show()







